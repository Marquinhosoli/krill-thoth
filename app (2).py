import io
import re
from copy import copy
from typing import Dict, List, Optional, Set, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="Krill → Thoth ERP", page_icon="📦", layout="wide")

VALID_STORES = {str(i) for i in range(1, 31)}
IGNORE_PRODUCT_NAMES = {"TOTAIS:", "TOTAL", "SUBTOTAL", "SUB-TOTAL", "PRODUTO", "PRODUTOS"}


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return " ".join(text.split())


def normalize_key(value) -> str:
    text = normalize_text(value).upper()
    text = re.sub(r"\s+", " ", text)
    return text


def find_order_header_row(df_raw: pd.DataFrame) -> int:
    for i in range(len(df_raw)):
        row = [normalize_text(v) for v in df_raw.iloc[i].tolist()]
        if "Loja" in row and "Descrição do Produto" in row and "Qtde." in row:
            return i
    raise ValueError("Não encontrei a linha de cabeçalho do pedido da Krill.")


def extract_order_table(order_file) -> pd.DataFrame:
    raw = pd.read_excel(order_file, header=None)
    header_row = find_order_header_row(raw)

    df = raw.iloc[header_row + 1:].copy()
    df.columns = raw.iloc[header_row]

    needed = ["Loja", "Descrição do Produto", "Qtde."]
    for col in needed:
        if col not in df.columns:
            raise ValueError(f"Coluna obrigatória ausente no pedido: {col}")

    df = df[needed].copy()
    df["Loja"] = df["Loja"].map(normalize_text)
    df["Descrição do Produto"] = df["Descrição do Produto"].map(normalize_text)
    df["Qtde."] = pd.to_numeric(df["Qtde."], errors="coerce").fillna(0)

    df = df[df["Descrição do Produto"] != ""]
    df = df[~df["Descrição do Produto"].map(normalize_key).isin(IGNORE_PRODUCT_NAMES)]
    df = df[df["Loja"].isin(VALID_STORES)]
    df = df[df["Loja"] != "0"]  # CD nunca entra
    df = df[df["Qtde."] > 0]

    if df.empty:
        raise ValueError("Nenhum item válido foi encontrado no pedido.")

    return df


def build_pivot(df: pd.DataFrame) -> pd.DataFrame:
    pivot = df.pivot_table(
        index="Descrição do Produto",
        columns="Loja",
        values="Qtde.",
        aggfunc="sum",
        fill_value=0,
    )
    ordered_cols = sorted([c for c in pivot.columns if str(c).isdigit()], key=lambda x: int(x))
    pivot = pivot.reindex(ordered_cols, axis=1)
    return pivot


def detect_model_layout(ws):
    product_header_row = None
    for row in range(1, min(ws.max_row, 10) + 1):
        if normalize_key(ws.cell(row, 1).value) in {"PRODUTO", "PRODUTOS"}:
            product_header_row = row
            break

    if not product_header_row:
        raise ValueError("Não encontrei a linha de produtos no modelo.")

    data_start_row = product_header_row + 1

    store_to_col = {}
    total_col = None
    cd_col = None

    max_scan_col = min(ws.max_column, 60)
    for col in range(2, max_scan_col + 1):
        top = normalize_key(ws.cell(1, col).value)
        second = normalize_text(ws.cell(product_header_row, col).value)

        if top and "TOTAL" in top:
            total_col = col
            continue

        if "CD" in top:
            cd_col = col
            continue

        store_num = ""
        if second.isdigit():
            store_num = second
        else:
            digits = "".join(ch for ch in top if ch.isdigit())
            if digits:
                store_num = digits

        if store_num and store_num != "0":
            store_to_col[store_num] = col

    if not store_to_col:
        raise ValueError("Não consegui localizar as colunas das lojas no modelo.")

    return {
        "product_header_row": product_header_row,
        "data_start_row": data_start_row,
        "store_to_col": store_to_col,
        "total_col": total_col,
        "cd_col": cd_col,
    }


def detect_model_products(ws, data_start_row: int) -> Tuple[Dict[str, int], List[str]]:
    key_to_row = {}
    ordered_products = []

    for row in range(data_start_row, ws.max_row + 1):
        product = normalize_text(ws.cell(row, 1).value)
        if not product:
            continue
        key = normalize_key(product)
        if key in IGNORE_PRODUCT_NAMES:
            continue
        ordered_products.append(product)
        key_to_row[key] = row

    return key_to_row, ordered_products


def copy_row_style(ws, source_row: int, target_row: int):
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        dst._style = copy(src._style)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def clear_model_values(ws, product_rows: Dict[str, int], store_cols: List[int], total_col: Optional[int], cd_col: Optional[int]):
    for row in product_rows.values():
        for col in store_cols:
            ws.cell(row, col).value = None
        if total_col:
            ws.cell(row, total_col).value = None
        if cd_col:
            ws.cell(row, cd_col).value = None


def split_products_strict(pivot: pd.DataFrame, frutas_model, legumes_model):
    frutas_key_to_row, frutas_names = frutas_model
    legumes_key_to_row, legumes_names = legumes_model

    frutas_keys = set(frutas_key_to_row.keys())
    legumes_keys = set(legumes_key_to_row.keys())

    frutas_items = []
    legumes_items = []
    unknown_items = []

    for product in pivot.index.tolist():
        key = normalize_key(product)
        if key in frutas_keys:
            frutas_items.append(product)
        elif key in legumes_keys:
            legumes_items.append(product)
        else:
            unknown_items.append(product)

    frutas_df = pivot.loc[frutas_items].copy() if frutas_items else pivot.iloc[0:0].copy()
    legumes_df = pivot.loc[legumes_items].copy() if legumes_items else pivot.iloc[0:0].copy()
    unknown_df = pivot.loc[unknown_items].copy() if unknown_items else pivot.iloc[0:0].copy()

    return frutas_df, legumes_df, unknown_df


def write_model_output(model_file, data: pd.DataFrame) -> bytes:
    wb = load_workbook(model_file)
    ws = wb.active

    layout = detect_model_layout(ws)
    product_rows, _ = detect_model_products(ws, layout["data_start_row"])

    clear_model_values(
        ws,
        product_rows,
        list(layout["store_to_col"].values()),
        layout["total_col"],
        layout["cd_col"],
    )

    used_keys = set()

    # Preenche itens existentes no modelo
    for source_product in data.index.tolist():
        product_key = normalize_key(source_product)
        if product_key not in product_rows:
            continue

        row = product_rows[product_key]
        used_keys.add(product_key)
        row_total = 0

        for store in data.columns:
            if store not in layout["store_to_col"]:
                continue
            qty = float(data.loc[source_product, store])
            if qty:
                ws.cell(row, layout["store_to_col"][store]).value = qty
                row_total += qty

        if layout["total_col"]:
            ws.cell(row, layout["total_col"]).value = row_total if row_total else None

        if layout["cd_col"]:
            ws.cell(row, layout["cd_col"]).value = None

    # Adiciona itens faltantes no final da planilha correspondente
    missing_products = [p for p in data.index.tolist() if normalize_key(p) not in used_keys]
    if missing_products:
        existing_rows = list(product_rows.values())
        style_source_row = max(existing_rows) if existing_rows else layout["data_start_row"]
        current_row = style_source_row + 1

        for product in missing_products:
            copy_row_style(ws, style_source_row, current_row)
            ws.cell(current_row, 1).value = product

            row_total = 0
            for store in data.columns:
                if store not in layout["store_to_col"]:
                    continue
                qty = float(data.loc[product, store])
                if qty:
                    ws.cell(current_row, layout["store_to_col"][store]).value = qty
                    row_total += qty

            if layout["total_col"]:
                ws.cell(current_row, layout["total_col"]).value = row_total if row_total else None

            if layout["cd_col"]:
                ws.cell(current_row, layout["cd_col"]).value = None

            current_row += 1

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def build_price_sheet(frutas: pd.DataFrame, legumes: pd.DataFrame) -> bytes:
    def make_df(data: pd.DataFrame) -> pd.DataFrame:
        if data.empty:
            return pd.DataFrame(columns=["CODIGO", "PRODUTO", "PRECO"])
        items = sorted(data.index.tolist())
        return pd.DataFrame({
            "CODIGO": [""] * len(items),
            "PRODUTO": items,
            "PRECO": [""] * len(items),
        })

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        make_df(frutas).to_excel(writer, sheet_name="FRUTAS", index=False)
        make_df(legumes).to_excel(writer, sheet_name="LEGUMES", index=False)
    out.seek(0)
    return out.getvalue()


def build_unknown_sheet(unknown: pd.DataFrame) -> bytes:
    out = io.BytesIO()
    if unknown.empty:
        df = pd.DataFrame(columns=["PRODUTO"])
    else:
        df = pd.DataFrame({"PRODUTO": unknown.index.tolist()}).sort_values("PRODUTO")
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="NAO_ENCONTRADOS", index=False)
    out.seek(0)
    return out.getvalue()


def metrics_for_display(df: pd.DataFrame) -> int:
    return 0 if df.empty else len(df.index.tolist())


def main():
    st.title("📦 Krill → Thoth ERP")
    st.caption("Versão rígida: respeita o modelo, ignora KRILL CD e não joga item desconhecido em categoria errada.")

    with st.sidebar:
        st.subheader("Como usar")
        st.write("1. Envie o pedido bruto da Krill.")
        st.write("2. Envie o modelo de FRUTAS.")
        st.write("3. Envie o modelo de LEGUMES.")
        st.write("4. Clique em Processar.")
        st.write("5. Baixe FRUTAS, LEGUMES e PREÇOS.")
        st.info("KRILL CD sempre fica vazio.")
        st.warning("Itens não encontrados em nenhum modelo ficam separados para conferência, sem contaminar FRUTAS ou LEGUMES.")

    c1, c2, c3 = st.columns(3)
    with c1:
        order_file = st.file_uploader("Pedido Krill", type=["xlsx", "xls"])
    with c2:
        frutas_model = st.file_uploader("Modelo FRUTAS", type=["xlsx", "xls"])
    with c3:
        legumes_model = st.file_uploader("Modelo LEGUMES", type=["xlsx", "xls"])

    process = st.button("PROCESSAR PEDIDO KRILL", type="primary", use_container_width=True)

    if process:
        if not order_file or not frutas_model or not legumes_model:
            st.error("Envie os 3 arquivos para continuar.")
            return

        try:
            order_df = extract_order_table(order_file)
            pivot = build_pivot(order_df)

            # Lê modelos uma vez para classificação segura
            frutas_wb = load_workbook(frutas_model)
            legumes_wb = load_workbook(legumes_model)
            frutas_ws = frutas_wb.active
            legumes_ws = legumes_wb.active

            frutas_layout = detect_model_layout(frutas_ws)
            legumes_layout = detect_model_layout(legumes_ws)

            frutas_model_info = detect_model_products(frutas_ws, frutas_layout["data_start_row"])
            legumes_model_info = detect_model_products(legumes_ws, legumes_layout["data_start_row"])

            # Reseta ponteiros dos uploads para reutilizar
            frutas_model.seek(0)
            legumes_model.seek(0)

            frutas_df, legumes_df, unknown_df = split_products_strict(
                pivot,
                frutas_model_info,
                legumes_model_info,
            )

            frutas_bytes = write_model_output(frutas_model, frutas_df)
            legumes_bytes = write_model_output(legumes_model, legumes_df)
            precos_bytes = build_price_sheet(frutas_df, legumes_df)
            unknown_bytes = build_unknown_sheet(unknown_df)

            st.success("Processamento concluído.")

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Itens FRUTAS", metrics_for_display(frutas_df))
            m2.metric("Itens LEGUMES", metrics_for_display(legumes_df))
            m3.metric("Lojas válidas", len(pivot.columns.tolist()))
            m4.metric("Não encontrados", metrics_for_display(unknown_df))

            if not unknown_df.empty:
                st.warning("Há itens fora dos modelos. Eles não foram jogados em FRUTAS ou LEGUMES por segurança.")

            d1, d2, d3 = st.columns(3)
            with d1:
                st.download_button(
                    "Baixar FRUTAS",
                    data=frutas_bytes,
                    file_name="KRILL_FRUTAS_Thoth.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with d2:
                st.download_button(
                    "Baixar LEGUMES",
                    data=legumes_bytes,
                    file_name="KRILL_LEGUMES_Thoth.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with d3:
                st.download_button(
                    "Baixar PREÇOS",
                    data=precos_bytes,
                    file_name="KRILL_PRECOS.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            if not unknown_df.empty:
                st.download_button(
                    "Baixar NÃO ENCONTRADOS",
                    data=unknown_bytes,
                    file_name="KRILL_NAO_ENCONTRADOS.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            with st.expander("Conferência do processamento"):
                lojas_texto = ", ".join([str(x) for x in pivot.columns.tolist()])
                st.write(f"**Lojas processadas:** {lojas_texto if lojas_texto else 'nenhuma'}")

                tab1, tab2, tab3 = st.tabs(["FRUTAS", "LEGUMES", "NÃO ENCONTRADOS"])

                with tab1:
                    if frutas_df.empty:
                        st.info("Nenhum item de fruta encontrado.")
                    else:
                        preview = frutas_df.reset_index().rename(columns={"index": "PRODUTO"})
                        st.dataframe(preview, use_container_width=True, hide_index=True)

                with tab2:
                    if legumes_df.empty:
                        st.info("Nenhum item de legume encontrado.")
                    else:
                        preview = legumes_df.reset_index().rename(columns={"index": "PRODUTO"})
                        st.dataframe(preview, use_container_width=True, hide_index=True)

                with tab3:
                    if unknown_df.empty:
                        st.success("Todos os itens bateram com algum modelo.")
                    else:
                        preview = pd.DataFrame({"PRODUTO": unknown_df.index.tolist()}).sort_values("PRODUTO")
                        st.dataframe(preview, use_container_width=True, hide_index=True)

        except Exception as e:
            st.error(f"Erro ao processar: {e}")


if __name__ == "__main__":
    main()
