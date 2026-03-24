from io import BytesIO
from pathlib import Path
from copy import copy
import re
import traceback

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="KRILL → THOTH (PRO)", page_icon="📦", layout="wide")

BASE_DIR = Path(__file__).resolve().parent
IGNORE_NAMES = {"", "TOTAL", "TOTAIS", "SUBTOTAL", "SUB-TOTAL", "PRODUTO", "PRODUTOS"}


def norm_text(v):
    if v is None:
        return ""
    t = str(v).strip()
    if t.lower() == "nan":
        return ""
    return " ".join(t.split())


def norm_key(v):
    return re.sub(r"\s+", " ", norm_text(v).upper())


def resolve_model_path(filename: str) -> Path:
    candidates = [
        BASE_DIR / filename,
        BASE_DIR / "models" / filename,
    ]
    for path in candidates:
        if path.exists():
            return path
    return BASE_DIR / filename


MODEL_FRUTAS = resolve_model_path("KRILL_FRUTAS_Branco (1).xlsx")
MODEL_LEGUMES = resolve_model_path("KRILL_LEGUMES_Branco (1).xlsx")


def find_header_row(raw: pd.DataFrame) -> int:
    for i in range(len(raw)):
        row = [norm_key(x) for x in raw.iloc[i].tolist()]

        has_loja = "LOJA" in row
        has_produto = (
            "DESCRIÇÃO DO PRODUTO" in row
            or "DESCRICAO DO PRODUTO" in row
            or "PRODUTO" in row
        )
        has_qtde = (
            "QTDE." in row
            or "QTDE" in row
            or "QUANTIDADE" in row
        )

        if has_loja and has_produto and has_qtde:
            return i

    raise ValueError(
        "Não encontrei o cabeçalho do pedido. "
        "Precisa existir Loja, Descrição do Produto e Qtde."
    )


def parse_price_series(series: pd.Series) -> pd.Series:
    if pd.api.types.is_numeric_dtype(series):
        return pd.to_numeric(series, errors="coerce")

    s = series.astype(str).str.strip()
    s = s.replace({"": None, "nan": None, "None": None, "<NA>": None})

    def convert(v):
        if v is None or pd.isna(v):
            return None
        # Limpa qualquer R$ ou espaço
        txt = str(v).strip().replace("R$", "").replace("r$", "").strip()

        if txt in ["", "nan", "<NA>", "None"]:
            return None

        # Arruma formatações com ponto e vírgula
        if "," in txt and "." in txt:
            if txt.rfind(",") > txt.rfind("."):
                txt = txt.replace(".", "").replace(",", ".")
            else:
                txt = txt.replace(",", "")
        elif "," in txt:
            txt = txt.replace(",", ".")

        try:
            return float(txt)
        except Exception:
            return None

    return s.map(convert)


def read_order(file):
    file_buffer = BytesIO(file.getvalue())
    raw = pd.read_excel(file_buffer, header=None)
    
    header_row = find_header_row(raw)

    row_vals = raw.iloc[header_row].tolist()
    new_cols = []
    seen = {}
    for c in row_vals:
        base_c = norm_key(c)
        if base_c == "":
            base_c = "VAZIO"
        if base_c in seen:
            seen[base_c] += 1
            new_cols.append(f"{base_c}_{seen[base_c]}")
        else:
            seen[base_c] = 0
            new_cols.append(base_c)

    df = raw.iloc[header_row + 1:].copy()
    df.columns = new_cols

    def get_exact_col(possibilities):
        for p in possibilities:
            k = norm_key(p)
            if k in df.columns:
                return df[k], k
        raise ValueError(f"Coluna obrigatória não encontrada. Procurado: {possibilities}")

    col_loja_s, debug_loja = get_exact_col(["LOJA"])
    col_produto_s, debug_produto = get_exact_col(["DESCRIÇÃO DO PRODUTO", "DESCRICAO DO PRODUTO", "PRODUTO"])
    col_qtde_s, debug_qtde = get_exact_col(["QTDE.", "QTDE", "QUANTIDADE"])

    # NOVO: Função super blindada que rejeita colunas bloqueadas
    def extract_coalesced(keyword_groups, exclude_words=[]):
        for group in keyword_groups:
            matching_cols = []
            for col in df.columns:
                base_col = re.sub(r'_\d+$', '', col).strip()
                if all(word in base_col for word in group):
                    # Garante que a coluna não contém as palavras bloqueadas (ex: "CX")
                    if not any(ex in base_col for ex in exclude_words):
                        matching_cols.append(col)
            
            if matching_cols:
                temp_df = df[matching_cols].copy()
                for c in temp_df.columns:
                    temp_df[c] = temp_df[c].apply(
                        lambda x: pd.NA if pd.isna(x) or str(x).strip() in ['', 'nan', 'None', '<NA>'] else x
                    )
                # Puxa o dado esteja ele em qual coluna mesclada estiver
                s = temp_df.bfill(axis=1).ffill(axis=1).iloc[:, 0]
                return s, re.sub(r'_\d+$', '', matching_cols[0])
                
        return None, "NÃO ENCONTRADA"

    # Procurando Código
    col_codigo_s, debug_codigo = extract_coalesced(
        [["GTIN"], ["PLU"], ["EAN"], ["BARRA"], ["REF", "FORNECEDOR"], ["CODIGO"], ["CÓDIGO"], ["COD"], ["CÓD"], ["ITEM"], ["SKU"]],
        exclude_words=["DESCRIÇÃO", "NOME"]
    )
    
    # Procurando Preço: PROÍBE TERMINANTEMENTE LER A CAIXA OU TOTAL
    col_preco_s, debug_preco = extract_coalesced(
        [
            ["CUSTO", "UN"], ["PRECO", "UN"], ["VALOR", "UN"],
            ["UNITARIO"], ["UNITÁRIO"], ["UNIT"], 
            ["CUSTO"], ["PRECO"], ["PREÇO"], ["VALOR"], ["VLR"]
        ],
        exclude_words=["CX", "CAIXA", "TOTAL", "SUBTOTAL", "BRUTO", "DESC", "FRETE", "IMPOSTO", "IPI", "ICMS"]
    )

    clean_df = pd.DataFrame()
    clean_df["Loja"] = col_loja_s
    clean_df["Descrição do Produto"] = col_produto_s
    clean_df["Qtde."] = col_qtde_s

    if col_codigo_s is not None:
        clean_df["CodigoPedido"] = col_codigo_s
    else:
        clean_df["CodigoPedido"] = ""

    if col_preco_s is not None:
        clean_df["PrecoPedido"] = col_preco_s
    else:
        clean_df["PrecoPedido"] = None

    clean_df["Loja"] = clean_df["Loja"].map(norm_text)
    clean_df["Descrição do Produto"] = clean_df["Descrição do Produto"].map(norm_text)
    clean_df["Qtde."] = pd.to_numeric(clean_df["Qtde."], errors="coerce").fillna(0)
    
    clean_df["CodigoPedido"] = clean_df["CodigoPedido"].map(lambda x: norm_text(x) if pd.notna(x) else "")
    clean_df["PrecoPedido"] = parse_price_series(clean_df["PrecoPedido"])

    clean_df = clean_df[clean_df["Descrição do Produto"] != ""]
    clean_df = clean_df[~clean_df["Descrição do Produto"].map(norm_key).isin(IGNORE_NAMES)]
    clean_df = clean_df[clean_df["Loja"].str.fullmatch(r"\d+")]
    clean_df = clean_df[clean_df["Loja"] != "0"]
    clean_df = clean_df[clean_df["Qtde."] > 0]

    if clean_df.empty:
        raise ValueError("Nenhum item válido foi encontrado no pedido.")

    debug_info = {
        "col_loja": debug_loja,
        "col_produto": debug_produto,
        "col_qtde": debug_qtde,
        "col_codigo": debug_codigo,
        "col_preco": debug_preco,
    }

    return clean_df, debug_info


def build_pivot(df: pd.DataFrame) -> pd.DataFrame:
    pivot = df.pivot_table(
        index="Descrição do Produto",
        columns="Loja",
        values="Qtde.",
        aggfunc="sum",
        fill_value=0,
    )
    cols = sorted([str(c) for c in pivot.columns], key=lambda x: int(x))
    pivot.columns = [str(c) for c in pivot.columns]
    return pivot.reindex(cols, axis=1)


def extract_store_number(v1, v2):
    c1 = norm_key(v1)
    c2 = norm_key(v2)

    for txt in (c1, c2):
        m = re.search(r"\bKRILL\s*(\d+)\b", txt)
        if m:
            return m.group(1)

    for txt in (c2, c1):
        if re.fullmatch(r"\d+", txt):
            return txt

    for txt in (c1, c2):
        numbers = re.findall(r"\d+", txt)
        if numbers:
            return numbers[0]

    return None


def model_map(ws):
    store_to_col = {}
    total_col = None
    cd_col = None

    for col in range(2, ws.max_column + 1):
        line1 = ws.cell(1, col).value
        line2 = ws.cell(2, col).value

        top = norm_key(line1)
        second = norm_key(line2)

        if "TOTAL" in top or "TOTAL" in second:
            total_col = col
            continue

        if "CD" in top or "CD" in second:
            cd_col = col
            continue

        loja = extract_store_number(line1, line2)
        if loja:
            store_to_col[loja] = col

    if not store_to_col:
        raise ValueError(
            "Não consegui mapear as lojas no modelo. "
            "Verifique se o cabeçalho contém KRILL X ou número da loja na linha 1 ou 2."
        )

    return store_to_col, total_col, cd_col


def product_rows(ws):
    rows = {}
    for row in range(3, ws.max_row + 1):
        prod = norm_text(ws.cell(row, 1).value)
        if prod and norm_key(prod) not in IGNORE_NAMES:
            rows[norm_key(prod)] = row
    return rows


@st.cache_data
def get_cached_product_rows(model_path_str: str):
    wb = load_workbook(model_path_str)
    return product_rows(wb.active)


def copy_row_style(ws, src_row, dst_row):
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        dst._style = copy(src._style)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def split_by_models(pivot, frutas_rows, legumes_rows):
    frutas_idx, legumes_idx, unknown_idx = [], [], []
    frutas_keys = set(frutas_rows.keys())
    legumes_keys = set(legumes_rows.keys())

    for prod in pivot.index.tolist():
        key = norm_key(prod)
        if key in frutas_keys:
            frutas_idx.append(prod)
        elif key in legumes_keys:
            legumes_idx.append(prod)
        else:
            unknown_idx.append(prod)

    empty = pivot.iloc[0:0].copy()
    frutas = pivot.loc[frutas_idx].copy() if frutas_idx else empty.copy()
    legumes = pivot.loc[legumes_idx].copy() if legumes_idx else empty.copy()
    unknown = pivot.loc[unknown_idx].copy() if unknown_idx else empty.copy()

    return frutas, legumes, unknown


def write_output(model_path: Path, data: pd.DataFrame) -> bytes:
    wb = load_workbook(str(model_path))
    ws = wb.active

    stores, total_col, cd_col = model_map(ws)
    prod_map = product_rows(ws)

    cols_to_clear = list(stores.values())
    if total_col: cols_to_clear.append(total_col)
    if cd_col: cols_to_clear.append(cd_col)

    for row in range(3, ws.max_row + 1):
        if norm_text(ws.cell(row, 1).value):
            for col in cols_to_clear:
                ws.cell(row, col).value = None

    used = set()

    for prod in data.index.tolist():
        key = norm_key(prod)
        if key not in prod_map:
            continue

        row = prod_map[key]
        used.add(key)
        row_total = 0

        for loja in data.columns:
            if loja in stores:
                val = float(data.loc[prod, loja])
                if val:
                    ws.cell(row, stores[loja]).value = val
                    row_total += val

        if total_col:
            ws.cell(row, total_col).value = row_total if row_total else None

        if cd_col:
            ws.cell(row, cd_col).value = None

    missing = [prod for prod in data.index.tolist() if norm_key(prod) not in used]

    if missing:
        last_filled = max(prod_map.values()) if prod_map else 3
        style_row = last_filled
        current_row = last_filled + 1

        for prod in missing:
            copy_row_style(ws, style_row, current_row)
            ws.cell(current_row, 1).value = prod

            row_total = 0
            for loja in data.columns:
                if loja in stores:
                    val = float(data.loc[prod, loja])
                    if val:
                        ws.cell(current_row, stores[loja]).value = val
                        row_total += val

            if total_col:
                ws.cell(current_row, total_col).value = row_total if row_total else None

            if cd_col:
                ws.cell(current_row, cd_col).value = None

            current_row += 1

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def build_prices(frutas: pd.DataFrame, legumes: pd.DataFrame, order_df: pd.DataFrame) -> bytes:
    base_precos = order_df[["Descrição do Produto", "CodigoPedido", "PrecoPedido"]].copy()
    
    base_precos["CodigoPedido"] = base_precos["CodigoPedido"].replace({"": pd.NA})
    base_precos["PrecoPedido"] = base_precos["PrecoPedido"].replace({"": pd.NA})
    
    base_precos = base_precos.sort_values(by=["PrecoPedido", "CodigoPedido"], na_position="last")
    base_precos = base_precos.drop_duplicates(subset=["Descrição do Produto"])
    base_precos["PRODUTO_KEY"] = base_precos["Descrição do Produto"].map(norm_key)

    def make_df(df):
        if df.empty:
            return pd.DataFrame(columns=["CÓDIGO", "PRODUTO", "PREÇO"])

        produtos = sorted(df.index.tolist(), key=lambda x: str(x).strip().upper())
        linhas = []

        for prod in produtos:
            key = norm_key(prod)
            achou = base_precos[base_precos["PRODUTO_KEY"] == key]

            if not achou.empty:
                row = achou.iloc[0]
                linhas.append(
                    {
                        "CÓDIGO": row["CodigoPedido"] if pd.notna(row["CodigoPedido"]) else "",
                        "PRODUTO": prod,
                        "PREÇO": row["PrecoPedido"] if pd.notna(row["PrecoPedido"]) else "",
                    }
                )
            else:
                linhas.append(
                    {
                        "CÓDIGO": "",
                        "PRODUTO": prod,
                        "PREÇO": "",
                    }
                )

        return pd.DataFrame(linhas)

    frutas_df = make_df(frutas)
    legumes_df = make_df(legumes)

    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        frutas_df.to_excel(writer, sheet_name="FRUTAS", index=False)
        legumes_df.to_excel(writer, sheet_name="LEGUMES", index=False)
        
        for sheet_name in ["FRUTAS", "LEGUMES"]:
            worksheet = writer.sheets[sheet_name]
            worksheet.column_dimensions['A'].width = 12
            worksheet.column_dimensions['B'].width = 45
            worksheet.column_dimensions['C'].width = 15

    out.seek(0)
    return out.getvalue()


def build_unknown(unknown: pd.DataFrame) -> bytes:
    out = BytesIO()

    if not unknown.empty:
        df = pd.DataFrame(
            {
                "PRODUTO": sorted(unknown.index.tolist(), key=lambda x: norm_key(x))
            }
        )
    else:
        df = pd.DataFrame(columns=["PRODUTO"])

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="NAO_ENCONTRADOS", index=False)

    out.seek(0)
    return out.getvalue()


st.title("📦 KRILL → THOTH (PRO)")
st.caption("Modelos fixos no sistema. Mapeamento por número da loja. KRILL CD sempre vazio.")

with st.sidebar:
    st.subheader("Como usar")
    st.write("1. Envie o pedido bruto da Krill.")
    st.write("2. Clique em Processar.")
    st.write("3. Baixe FRUTAS, LEGUMES, PREÇOS e NÃO ENCONTRADOS.")
    st.info("O sistema busca os modelos automaticamente na raiz do app ou na pasta models.")
    st.info("A planilha de PREÇOS usa os códigos e preços do próprio pedido atual.")
    st.info("Regra: pedido loja X = coluna KRILL X.")

uploaded = st.file_uploader("Pedido Krill", type=["xlsx", "xls"])

if st.button("PROCESSAR", use_container_width=True, type="primary"):
    if not uploaded:
        st.error("Envie o pedido para continuar.")
    else:
        try:
            if not MODEL_FRUTAS.exists():
                st.error(f"Modelo FRUTAS não encontrado: {MODEL_FRUTAS.name}. Verifique se o arquivo está na pasta.")
                st.stop()

            if not MODEL_LEGUMES.exists():
                st.error(f"Modelo LEGUMES não encontrado: {MODEL_LEGUMES.name}. Verifique se o arquivo está na pasta.")
                st.stop()

            order_df, debug_info = read_order(uploaded)
            pivot = build_pivot(order_df)

            frutas_rows = get_cached_product_rows(str(MODEL_FRUTAS))
            legumes_rows = get_cached_product_rows(str(MODEL_LEGUMES))

            frutas_df, legumes_df, unknown_df = split_by_models(
                pivot, frutas_rows, legumes_rows
            )

            frutas_file = write_output(MODEL_FRUTAS, frutas_df)
            legumes_file = write_output(MODEL_LEGUMES, legumes_df)
            prices_file = build_prices(frutas_df, legumes_df, order_df)
            unknown_file = build_unknown(unknown_df)

            st.success("Processamento concluído com sucesso.")

            st.caption(
                f"Colunas encontradas no pedido → "
                f"Loja: {debug_info['col_loja']} | "
                f"Produto: {debug_info['col_produto']} | "
                f"Qtde: {debug_info['col_qtde']} | "
                f"Código: {debug_info['col_codigo']} | "
                f"Preço: {debug_info['col_preco']}"
            )

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Itens FRUTAS", len(frutas_df.index))
            c2.metric("Itens LEGUMES", len(legumes_df.index))
            c3.metric("Lojas processadas", len(pivot.columns))
            c4.metric("Não encontrados", len(unknown_df.index))

            d1, d2, d3 = st.columns(3)
            with d1:
                st.download_button(
                    "Baixar FRUTAS",
                    frutas_file,
                    "KRILL_FRUTAS_Thoth.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with d2:
                st.download_button(
                    "Baixar LEGUMES",
                    legumes_file,
                    "KRILL_LEGUMES_Thoth.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with d3:
                st.download_button(
                    "Baixar PREÇOS",
                    prices_file,
                    "KRILL_PRECOS.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            if not unknown_df.empty:
                st.download_button(
                    "Baixar NÃO ENCONTRADOS",
                    unknown_file,
                    "KRILL_NAO_ENCONTRADOS.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

        except Exception as e:
            st.error(f"Erro ao processar: {e}")
            with st.expander("Ver detalhes do erro (Envie isso para suporte)"):
                st.code(traceback.format_exc())
